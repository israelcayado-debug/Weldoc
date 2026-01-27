from datetime import timedelta
from pathlib import Path

from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import extend_schema

from apps.welds import models as weld_models
from apps.projects import models as project_models
from apps.wpq import models as wpq_models
from apps.wps import models as wps_models
from . import models
from . import serializers


class ProgressReportView(APIView):
    serializer_class = serializers.ProgressReportSerializer

    @extend_schema(responses=serializers.ProgressReportSerializer)
    def get(self, request):
        project_id = request.query_params.get("project_id")
        if not project_id:
            return Response({"code": "missing_project", "message": "project_id requerido."}, status=400)
        qs = weld_models.Weld.objects.filter(project_id=project_id)
        total = qs.count()
        by_status = {
            "planned": qs.filter(status="planned").count(),
            "in_progress": qs.filter(status="in_progress").count(),
            "completed": qs.filter(status="completed").count(),
            "repair": qs.filter(status="repair").count(),
        }
        return Response({"project_id": project_id, "total_welds": total, "by_status": by_status})


class ExpiryReportView(APIView):
    serializer_class = serializers.ExpiryReportSerializer

    @extend_schema(responses=serializers.ExpiryReportSerializer)
    def get(self, request):
        project_id = request.query_params.get("project_id")
        if not project_id:
            return Response({"code": "missing_project", "message": "project_id requerido."}, status=400)
        today = timezone.now().date()
        warn_date = today + timedelta(days=30)
        continuity = wpq_models.WelderContinuity.objects.filter(
            welder__continuitylog__weld__project_id=project_id
        ).distinct()
        expiring = continuity.filter(continuity_due_date__lte=warn_date, status="in_continuity")
        out = continuity.filter(status="out_of_continuity")
        expiring_list = [
            {
                "welder_id": str(c.welder_id),
                "name": c.welder.name,
                "due_date": c.continuity_due_date,
            }
            for c in expiring
        ]
        out_list = [
            {
                "welder_id": str(c.welder_id),
                "name": c.welder.name,
                "last_activity": c.last_activity_date,
            }
            for c in out
        ]
        return Response(
            {
                "project_id": project_id,
                "expiring_30_days": expiring_list,
                "out_of_continuity": out_list,
            }
        )


class ExportWeldingListView(APIView):
    serializer_class = serializers.ExportWeldingListRequestSerializer

    @extend_schema(
        request=serializers.ExportWeldingListRequestSerializer,
        responses=serializers.ExportStatusSerializer,
    )
    def post(self, request):
        project_id = request.data.get("project_id")
        status_filter = request.data.get("status")
        drawing_id = request.data.get("drawing_id")
        date_from = request.data.get("date_from")
        date_to = request.data.get("date_to")
        if not project_id:
            return Response({"code": "missing_project", "message": "project_id requerido."}, status=400)
        report = models.Report.objects.create(
            project_id=project_id,
            type="welding_list",
            params_json={
                "filters": request.data.get("filters", {}),
                "status": status_filter,
                "drawing_id": drawing_id,
                "date_from": date_from,
                "date_to": date_to,
            },
        )
        welds = weld_models.Weld.objects.filter(project_id=project_id)
        if status_filter:
            welds = welds.filter(status=status_filter)
        if drawing_id:
            welds = welds.filter(drawing_id=drawing_id)
        if date_from:
            welds = welds.filter(
                Q(closed_at__date__gte=date_from)
                | Q(closed_at__isnull=True, status="planned")
            )
        if date_to:
            welds = welds.filter(
                Q(closed_at__date__lte=date_to)
                | Q(closed_at__isnull=True, status="planned")
            )
        wb = Workbook()
        ws = wb.active
        ws.title = "Welding List"
        ws.append(["number", "status", "drawing", "closed_at"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for w in welds:
            ws.append(
                [
                    w.number,
                    w.status,
                    w.drawing.code if w.drawing else "",
                    w.closed_at.isoformat() if w.closed_at else ("planned" if w.status == "planned" else ""),
                ]
            )
        report.file_path = f"welding_list_{report.id}.xlsx"
        report.params_json["format"] = "xlsx"
        report.save(update_fields=["file_path", "params_json"])
        path = Path(settings.MEDIA_ROOT) / report.file_path
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return Response({"export_id": str(report.id), "status": "ready", "file_path": report.file_path})


class ExportQualificationsView(APIView):
    serializer_class = serializers.ExportQualificationsRequestSerializer

    @extend_schema(
        request=serializers.ExportQualificationsRequestSerializer,
        responses=serializers.ExportStatusSerializer,
    )
    def post(self, request):
        project_id = request.data.get("project_id")
        export_type = request.data.get("type")
        if not project_id or not export_type:
            return Response({"code": "missing_params", "message": "project_id y type requeridos."}, status=400)
        report = models.Report.objects.create(
            project_id=project_id, type=f"qual_{export_type}", params_json={}
        )
        wb = Workbook()
        ws = wb.active
        ws.title = f"Qualifications {export_type}"
        ws.append(["code", "status"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        if export_type == "WPS":
            for wps in wps_models.Wps.objects.filter(project_id=project_id):
                ws.append([wps.code, wps.status])
        elif export_type == "WPQ":
            for wpq in wpq_models.Wpq.objects.all():
                ws.append([wpq.code, wpq.status])
        report.file_path = f"qual_{export_type}_{report.id}.xlsx"
        report.params_json["format"] = "xlsx"
        report.save(update_fields=["file_path", "params_json"])
        path = Path(settings.MEDIA_ROOT) / report.file_path
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return Response({"export_id": str(report.id), "status": "ready", "file_path": report.file_path})


class ExportDossierView(APIView):
    serializer_class = serializers.ExportDossierRequestSerializer

    @extend_schema(
        request=serializers.ExportDossierRequestSerializer,
        responses=serializers.ExportStatusSerializer,
    )
    def post(self, request):
        project_id = request.data.get("project_id")
        include = request.data.get("include", [])
        if not project_id:
            return Response({"code": "missing_project", "message": "project_id requerido."}, status=400)
        dossier = models.Dossier.objects.create(
            project_id=project_id, config_json={"include": include}
        )
        project = project_models.Project.objects.select_related("client").filter(id=project_id).first()
        project_name = project.name if project else None
        project_number = project.code if project else ""
        project_client = project.client.name if project and project.client else ""
        project_po = (project.purchase_order or "") if project else ""
        file_name = f"dossier_{dossier.id}.pdf"
        dossier.file_path = file_name
        dossier.save(update_fields=["file_path"])

        path = Path(settings.MEDIA_ROOT) / file_name
        path.parent.mkdir(parents=True, exist_ok=True)

        pdf = canvas.Canvas(str(path), pagesize=A4)
        width, height = A4
        page_num = 1

        def draw_header_footer():
            pdf.setFont("Helvetica-Bold", 11)
            pdf.drawString(2 * cm, height - 1.5 * cm, "Welding Dossier")
            pdf.setFont("Helvetica", 9)
            pdf.drawString(
                2 * cm,
                height - 2.1 * cm,
                f"Project: {project_name or 'Project'} ({project_number})",
            )
            pdf.drawString(
                2 * cm,
                height - 2.7 * cm,
                f"Customer: {project_client} | PO: {project_po}",
            )
            pdf.drawRightString(width - 2 * cm, height - 1.5 * cm, timezone.now().date().isoformat())
            pdf.setFont("Helvetica", 9)
            pdf.drawRightString(width - 2 * cm, 1.2 * cm, f"Page {page_num}")

        draw_header_footer()
        y = height - 3 * cm
        pdf.setFont("Helvetica", 12)
        pdf.drawString(2 * cm, y, f"Export ID: {dossier.id}")
        y -= 1.0 * cm

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(2 * cm, y, "Contents")
        y -= 0.6 * cm
        pdf.setFont("Helvetica", 11)
        items = include or [
            "welding_list",
            "qualifications",
            "inspections",
            "materials",
            "consumables",
        ]
        for entry in items:
            if y < 3 * cm:
                pdf.showPage()
                page_num += 1
                draw_header_footer()
                y = height - 3 * cm
            pdf.drawString(2.5 * cm, y, f"- {entry}")
            y -= 0.5 * cm

        pdf.showPage()
        page_num += 1
        draw_header_footer()
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(2 * cm, height - 3 * cm, "Summary")
        pdf.setFont("Helvetica", 11)
        total_welds = weld_models.Weld.objects.filter(project_id=project_id).count()
        pdf.drawString(2 * cm, height - 4 * cm, f"Total welds: {total_welds}")

        def add_section(title, headers, rows):
            nonlocal y
            pdf.showPage()
            nonlocal page_num
            page_num += 1
            draw_header_footer()
            y = height - 3 * cm
            pdf.setFont("Helvetica-Bold", 12)
            pdf.drawString(2 * cm, y, title)
            y -= 0.7 * cm
            pdf.setFont("Helvetica-Bold", 10)
            x = 2 * cm
            for header in headers:
                pdf.drawString(x, y, header)
                x += 5 * cm
            y -= 0.5 * cm
            pdf.setFont("Helvetica", 10)
            for row in rows:
                if y < 2 * cm:
                    pdf.showPage()
                    page_num += 1
                    draw_header_footer()
                    y = height - 3 * cm
                x = 2 * cm
                for value in row:
                    pdf.drawString(x, y, str(value))
                    x += 5 * cm
                y -= 0.45 * cm

        if not include or "welding_list" in include:
            weld_rows = [
                (w.number, w.status, w.drawing.code if w.drawing else "")
                for w in weld_models.Weld.objects.filter(project_id=project_id).select_related("drawing")
            ]
            add_section("Welding List", ["Number", "Status", "Drawing"], weld_rows)

        if not include or "qualifications" in include:
            wps_rows = [
                (w.code, w.status, w.standard)
                for w in wps_models.Wps.objects.filter(project_id=project_id)
            ]
            add_section("WPS", ["Code", "Status", "Standard"], wps_rows)
            wpq_rows = [
                (w.code, w.status, w.welder.name)
                for w in wpq_models.Wpq.objects.select_related("welder")
            ]
            add_section("WPQ", ["Code", "Status", "Welder"], wpq_rows)

        if not include or "inspections" in include:
            insp_rows = [
                (i.weld.number, i.stage, i.result)
                for i in weld_models.VisualInspection.objects.select_related("weld").filter(
                    weld__project_id=project_id
                )
            ]
            add_section("Visual Inspections", ["Weld", "Stage", "Result"], insp_rows)

        if not include or "materials" in include:
            material_rows = [
                (m.weld.number, m.material.spec, m.heat_number)
                for m in weld_models.WeldMaterial.objects.select_related("weld", "material").filter(
                    weld__project_id=project_id
                )
            ]
            add_section("Materials", ["Weld", "Material", "Heat"], material_rows)

        if not include or "consumables" in include:
            consumable_rows = [
                (c.weld.number, c.consumable.spec, c.batch or "")
                for c in weld_models.WeldConsumable.objects.select_related("weld", "consumable").filter(
                    weld__project_id=project_id
                )
            ]
            add_section("Consumables", ["Weld", "Consumable", "Batch"], consumable_rows)
        pdf.save()

        return Response({"export_id": str(dossier.id), "status": "ready", "file_path": dossier.file_path})


class ExportStatusView(APIView):
    serializer_class = serializers.ExportStatusSerializer

    @extend_schema(responses=serializers.ExportStatusSerializer)
    def get(self, request, export_id):
        report = models.Report.objects.filter(id=export_id).first()
        if report:
            status_value = "ready" if report.file_path else "queued"
            return Response({"export_id": str(report.id), "status": status_value, "file_path": report.file_path})
        dossier = models.Dossier.objects.filter(id=export_id).first()
        if dossier:
            status_value = "ready" if dossier.file_path else "queued"
            return Response({"export_id": str(dossier.id), "status": status_value, "file_path": dossier.file_path})
        return Response({"code": "not_found", "message": "export_id no existe."}, status=404)


class ExportDownloadView(APIView):
    serializer_class = serializers.ExportStatusSerializer

    @extend_schema(responses=OpenApiTypes.BINARY)
    def get(self, request, export_id):
        report = models.Report.objects.filter(id=export_id).first()
        if report and report.file_path:
            path = Path(settings.MEDIA_ROOT) / report.file_path
            if path.exists():
                if report.file_path.lower().endswith(".xlsx"):
                    resp = HttpResponse(
                        path.read_bytes(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                elif report.file_path.lower().endswith(".pdf"):
                    resp = HttpResponse(path.read_bytes(), content_type="application/pdf")
                else:
                    resp = HttpResponse(path.read_text(encoding="utf-8"), content_type="text/csv")
                resp["Content-Disposition"] = f"attachment; filename={report.file_path}"
                return resp
        dossier = models.Dossier.objects.filter(id=export_id).first()
        if dossier and dossier.file_path:
            path = Path(settings.MEDIA_ROOT) / dossier.file_path
            if path.exists():
                resp = HttpResponse(path.read_bytes(), content_type="application/pdf")
                resp["Content-Disposition"] = f"attachment; filename={dossier.file_path}"
                return resp
        return Response({"code": "not_found", "message": "file not available."}, status=404)
