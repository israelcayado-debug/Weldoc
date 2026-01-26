from pathlib import Path

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

from apps.welds import models as weld_models
from apps.wps import models as wps_models
from apps.wpq import models as wpq_models
from apps.projects import models as project_models
from . import models


def _export_welding_list(project_id, filters=None, status_filter=None):
    report = models.Report.objects.create(
        project_id=project_id,
        type="welding_list",
        params_json={
            "filters": filters or {},
            "status": status_filter,
        },
    )
    welds = weld_models.Weld.objects.filter(project_id=project_id)
    drawing_id = (filters or {}).get("drawing_id")
    date_from = (filters or {}).get("date_from")
    date_to = (filters or {}).get("date_to")
    if status_filter:
        welds = welds.filter(status=status_filter)
    if drawing_id:
        welds = welds.filter(drawing_id=drawing_id)
    if date_from:
        welds = welds.filter(
            Q(closed_at__date__gte=date_from)
            | Q(closed_at__isnull=True, status="planned")
        )
    if date_to:
        welds = welds.filter(
            Q(closed_at__date__lte=date_to)
            | Q(closed_at__isnull=True, status="planned")
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Welding List"
    ws.append(["number", "status", "drawing", "closed_at"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for weld in welds:
        ws.append(
            [
                weld.number,
                weld.status,
                weld.drawing.code if weld.drawing else "",
                weld.closed_at.isoformat()
                if weld.closed_at
                else ("planned" if weld.status == "planned" else ""),
            ]
        )
    report.file_path = f"welding_list_{report.id}.xlsx"
    report.params_json["format"] = "xlsx"
    report.save(update_fields=["file_path", "params_json"])
    path = Path(settings.MEDIA_ROOT) / report.file_path
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return report


def _export_qualifications(project_id, export_type):
    report = models.Report.objects.create(
        project_id=project_id, type=f"qual_{export_type}", params_json={}
    )
    wb = Workbook()
    ws = wb.active
    ws.title = f"Qualifications {export_type}"
    ws.append(["code", "status"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if export_type == "WPS":
        for wps in wps_models.Wps.objects.filter(project_id=project_id):
            ws.append([wps.code, wps.status])
    elif export_type == "WPQ":
        for wpq in wpq_models.Wpq.objects.all():
            ws.append([wpq.code, wpq.status])
    report.file_path = f"qual_{export_type}_{report.id}.xlsx"
    report.params_json["format"] = "xlsx"
    report.save(update_fields=["file_path", "params_json"])
    path = Path(settings.MEDIA_ROOT) / report.file_path
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return report


def _export_dossier(project_id, include):
    dossier = models.Dossier.objects.create(
        project_id=project_id, config_json={"include": include}
    )
    project_name = (
        weld_models.Weld.objects.filter(project_id=project_id)
        .values_list("project__name", flat=True)
        .first()
    )
    file_name = f"dossier_{dossier.id}.pdf"
    dossier.file_path = file_name
    dossier.save(update_fields=["file_path"])

    path = Path(settings.MEDIA_ROOT) / file_name
    path.parent.mkdir(parents=True, exist_ok=True)

    pdf = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    page_num = 1

    def draw_header_footer():
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(2 * cm, height - 1.5 * cm, "Dossier de Soldadura")
        pdf.setFont("Helvetica", 9)
        pdf.drawString(2 * cm, height - 2.1 * cm, f"Proyecto: {project_name or 'Proyecto'}")
        pdf.drawRightString(width - 2 * cm, height - 1.5 * cm, timezone.now().date().isoformat())
        pdf.setFont("Helvetica", 9)
        pdf.drawRightString(width - 2 * cm, 1.2 * cm, f"Pagina {page_num}")

    draw_header_footer()
    y = height - 3 * cm
    pdf.setFont("Helvetica", 12)
    pdf.drawString(2 * cm, y, f"Export ID: {dossier.id}")
    y -= 1.0 * cm

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(2 * cm, y, "Contenido")
    y -= 0.6 * cm
    pdf.setFont("Helvetica", 11)
    items = include or [
        "welding_list",
        "qualifications",
        "inspections",
        "materials",
        "consumables",
    ]
    for entry in items:
        if y < 3 * cm:
            pdf.showPage()
            page_num += 1
            draw_header_footer()
            y = height - 3 * cm
        pdf.drawString(2.5 * cm, y, f"- {entry}")
        y -= 0.5 * cm

    pdf.showPage()
    page_num += 1
    draw_header_footer()
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(2 * cm, height - 3 * cm, "Resumen")
    pdf.setFont("Helvetica", 11)
    total_welds = weld_models.Weld.objects.filter(project_id=project_id).count()
    pdf.drawString(2 * cm, height - 4 * cm, f"Total de soldaduras: {total_welds}")

    def add_section(title, headers, rows):
        nonlocal y
        pdf.showPage()
        nonlocal page_num
        page_num += 1
        draw_header_footer()
        y = height - 3 * cm
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(2 * cm, y, title)
        y -= 0.7 * cm
        pdf.setFont("Helvetica-Bold", 10)
        x = 2 * cm
        for header in headers:
            pdf.drawString(x, y, header)
            x += 5 * cm
        y -= 0.5 * cm
        pdf.setFont("Helvetica", 10)
        for row in rows:
            if y < 2 * cm:
                pdf.showPage()
                page_num += 1
                draw_header_footer()
                y = height - 3 * cm
            x = 2 * cm
            for value in row:
                pdf.drawString(x, y, str(value))
                x += 5 * cm
            y -= 0.45 * cm

    if not include or "welding_list" in include:
        weld_rows = [
            (w.number, w.status, w.drawing.code if w.drawing else "")
            for w in weld_models.Weld.objects.filter(project_id=project_id).select_related("drawing")
        ]
        add_section("Welding List", ["Numero", "Estado", "Plano"], weld_rows)

    if not include or "qualifications" in include:
        wps_rows = [
            (w.code, w.status, w.standard)
            for w in wps_models.Wps.objects.filter(project_id=project_id)
        ]
        add_section("WPS", ["Codigo", "Estado", "Norma"], wps_rows)
        wpq_rows = [
            (w.code, w.status, w.welder.name)
            for w in wpq_models.Wpq.objects.select_related("welder")
        ]
        add_section("WPQ", ["Codigo", "Estado", "Soldador"], wpq_rows)

    if not include or "inspections" in include:
        insp_rows = [
            (i.weld.number, i.stage, i.result)
            for i in weld_models.VisualInspection.objects.select_related("weld").filter(
                weld__project_id=project_id
            )
        ]
        add_section("Inspecciones Visuales", ["Soldadura", "Etapa", "Resultado"], insp_rows)

    if not include or "materials" in include:
        material_rows = [
            (m.weld.number, m.material.spec, m.heat_number)
            for m in weld_models.WeldMaterial.objects.select_related("weld", "material").filter(
                weld__project_id=project_id
            )
        ]
        add_section("Materiales", ["Soldadura", "Material", "Heat"], material_rows)

    if not include or "consumables" in include:
        consumable_rows = [
            (c.weld.number, c.consumable.spec, c.batch or "")
            for c in weld_models.WeldConsumable.objects.select_related("weld", "consumable").filter(
                weld__project_id=project_id
            )
        ]
        add_section("Consumibles", ["Soldadura", "Consumible", "Batch"], consumable_rows)

    pdf.save()
    return dossier


@login_required
def exports(request):
    message = None
    export_id = None
    export_file = None
    selected_project = None
    selected_status = None
    if request.method == "POST":
        action = request.POST.get("action")
        project_id = request.POST.get("project_id")
        selected_project = project_id
        if not project_id:
            message = "project_id requerido."
        elif action == "welding_list":
            status_filter = request.POST.get("status")
            drawing_id = request.POST.get("drawing_id")
            date_from = request.POST.get("date_from")
            date_to = request.POST.get("date_to")
            selected_status = status_filter
            report = _export_welding_list(
                project_id,
                status_filter=status_filter,
                filters={
                    "drawing_id": drawing_id,
                    "date_from": date_from,
                    "date_to": date_to,
                },
            )
            export_id = report.id
            export_file = report.file_path
        elif action == "qualifications":
            export_type = request.POST.get("qual_type")
            if export_type not in ("WPS", "WPQ"):
                message = "Tipo invalido."
            else:
                report = _export_qualifications(project_id, export_type)
                export_id = report.id
                export_file = report.file_path
        elif action == "dossier":
            include = request.POST.getlist("include")
            dossier = _export_dossier(project_id, include)
            export_id = dossier.id
            export_file = dossier.file_path
        else:
            message = "Accion invalida."
    projects = project_models.Project.objects.all().order_by("name")
    return render(
        request,
        "reports/exports.html",
        {
            "message": message,
            "export_id": export_id,
            "export_file": export_file,
            "projects": projects,
            "selected_project": selected_project,
            "selected_status": selected_status,
            "filters": request.POST if request.method == "POST" else {},
        },
    )


@login_required
def export_detail(request, export_id):
    report = models.Report.objects.filter(id=export_id).first()
    dossier = None
    if not report:
        dossier = models.Dossier.objects.filter(id=export_id).first()
    item = report or dossier
    return render(request, "reports/export_detail.html", {"item": item})


@login_required
def export_history(request):
    reports = models.Report.objects.select_related("project").all().order_by("-id")
    dossiers = models.Dossier.objects.select_related("project").all().order_by("-id")
    return render(
        request,
        "reports/history.html",
        {"reports": reports, "dossiers": dossiers},
    )
